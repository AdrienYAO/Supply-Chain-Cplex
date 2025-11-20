# -*- coding: utf-8 -*-
"""
MainPrincipale.py
-----------------
Script principal pour :
1. Générer les données avec Generateur_Donnees.py
2. Lancer l'optimisation CMS via CMS_Optimization
3. Extraire les résultats et les sauvegarder dans Excel
4. Traiter le fichier .ltf
5. Ajouter un onglet de synthèse des résultats.
"""

import subprocess
import os
import sys
import re
import statistics
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# --- Chemins des fichiers ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PYTHON_CPLEX = sys.executable  # Utiliser le Python actuel
GENERATOR = os.path.join(BASE_DIR, "Generateur_Donnees.py")
TXT_FILE = os.path.join(BASE_DIR, "resultats_optimisation.txt")
LTF_FILE = os.path.join(BASE_DIR, "DataFinal.ltf")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "resultats_complet.xlsx")

# --- En-têtes Excel ---
headers_txt = [
    "Nom fichier", "SCENARIO", "INSTANCE SCENARIO",
    "d11","d12","d21","d22","d31","d32","d41","d42",
    "LP","CL","OBJ.VALUE","T.PROD (X)","T.REPORT(B)","BINARY (Z)",
    "T.SOUS-T (Y)","T.ACH-M(NAJ)","T.VENT-M(NRE)","PRECESS TIME",
    "T.NONZERO","T.VARIABLES","T.CONTRAINTES","ITERATIONS"
]

headers_ltf = [
    "Nom fichier", "SCENARIO", "INSTANCE SCENARIO",
    "d11","d12","d21","d22","d31","d32","d41","d42",
    "LP","CL","T.DEMAND","OBJ.VALUE","OBJ.BOUND","T.PROD (X)","T.REPORT(B)","BINARY (Z)",
    "T.SOUS-T (Y)","T.ACH-M(NAJ)","T.VENT-M(NRE)","PRECESS TIME","T.INTEGER","T.NONZERO",
    "T.VARIABLES","T.CONTRAINTES","V.CONTRAINTES","ITERATIONS"
]

# --- Étape 1 : Génération des données ---
print("Exécution de Generateur_Donnees.py avec cplex-env...")
try:
    subprocess.run([PYTHON_CPLEX, GENERATOR], check=True)
    print("[OK] Generateur_Donnees.py exécuté avec succès !\n")
except subprocess.CalledProcessError as e:
    print("[ERREUR] Erreur pendant l’exécution du générateur :")
    print(e)
    sys.exit(1)

# --- Étape 2 : Lancement de l’optimisation CMS ---
print("Lancement de l'optimisation CMS via CMS_Optimization...")
try:
    from CMS_Optimization import run_cms_optimization
    run_cms_optimization()
    print("[OK] Optimisation CMS terminée avec succès !\n")
except Exception as e:
    print("[ERREUR] Erreur pendant l’exécution de l’optimisation CMS :")
    print(e)
    sys.exit(1)
    
from run_lingo import run_lingo_model

print("===== Étape 2 : Lancement du modèle LINGO =====")
run_lingo_model()
print("===== Modèle LINGO exécuté avec succès =====\n")


# --- Création / chargement du fichier Excel ---
if os.path.exists(OUTPUT_EXCEL):
    wb = load_workbook(OUTPUT_EXCEL)
else:
    wb = Workbook()

def create_sheet(workbook, sheet_name, headers):
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(sheet_name)
        ws.append(headers)
    return ws

ws_txt = create_sheet(wb, "TXT_Resultats", headers_txt)
ws_ltf = create_sheet(wb, "LTF_Resultats", headers_ltf)

# --- Déterminer SCENARIO et INSTANCE ---
def next_scenario_instance(ws):
    max_instance_scenario = 0
    last_scenario = 0
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        scenario_val = row[1]
        instance_val = row[2]
        if scenario_val is None or instance_val is None:
            continue
        if scenario_val > last_scenario:
            last_scenario = scenario_val
            max_instance_scenario = instance_val
        elif scenario_val == last_scenario and instance_val > max_instance_scenario:
            max_instance_scenario = instance_val
    if max_instance_scenario >= 50:
        return last_scenario + 1, 1
    else:
        return last_scenario, max_instance_scenario + 1

scenario, instance_scenario = next_scenario_instance(ws_txt)

def process_txt(file_path):
    Demande_values = {}
    lp_value = cl_value = ""
    total_nonzeros = 0
    sum_X = sum_Y = sum_B = sum_NAJ = sum_NRE = 0
    iterations = None  # valeur par défaut
    specific_results = {}
    tol = 1e-6

    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    in_dem_section = False
    for line in lines:
        line = line.strip()
        
        # Section Demande
        if line.startswith("Demande :"):
            in_dem_section = True
            continue
        if in_dem_section:
            if line == "" or not line.startswith("("):
                in_dem_section = False
            else:
                match = re.search(r"\((\d+),\s*(\d+)\)\s+([\d\.]+)", line)
                if match:
                    key = f"d{match.group(1)}{match.group(2)}"
                    Demande_values[key] = int(round(float(match.group(3))))

        # Variables X,Y,B,NAJ,NRE
        for var_type, regex in [
            ("X", r"X\(([\d,\s]+)\)\s*=\s*([\d\.]+)"),
            ("Y", r"Y\(([\d,\s]+)\)\s*=\s*([\d\.]+)"),
            ("B", r"B\(([\d,\s]+)\)\s*=\s*([\d\.]+)"),
            ("NAJ", r"NAJ\(([\d,\s]+)\)\s*=\s*([\d\.]+)"),
            ("NRE", r"NRE\(([\d,\s]+)\)\s*=\s*([\d\.]+)")
        ]:
            m = re.search(regex, line)
            if m:
                val = float(m.group(2))
                if abs(val) > tol:
                    if var_type == "X": sum_X += val
                    elif var_type == "Y": sum_Y += val
                    elif var_type == "B": sum_B += val
                    elif var_type == "NAJ": sum_NAJ += val
                    elif var_type == "NRE": sum_NRE += val

        # Patterns pour autres informations
        patterns = {
            "Fonction objectif": r"Fonction objectif\s*[:=]\s*([-+]?\d*\.\d+|\d+)",
            "Variables totales": r"Variables totales\s*:\s*(\d+)",
            "Nombre d'itérations": r"(?i)Nombre\s+d'itérations\s*[:=]?\s*(\d+)",  # <- robust
            "Contraintes totales": r"Contraintes totales\s*:\s*(\d+)",
            "Temps d'exécution": r"Temps d'exécution\s*[:=]?\s*(\d*\.\d+|\d+)",
            "nonzeros": r"nonzeros\s*[:=]?\s*(\d+)",
            "LP": r"LP\s*[:=]\s*([\d\.]+)",
            "CL": r"CL\s*[:=]\s*([\d\.]+)"
        }

        for key, pattern in patterns.items():
            m = re.search(pattern, line)
            if m:
                val = m.group(1)
                if key == "LP": 
                    lp_value = val
                elif key == "CL": 
                    cl_value = val
                elif key == "nonzeros": 
                    total_nonzeros = int(val)
                elif key.lower() == "nombre d'itérations":  # récupérer iterations
                    iterations = int(val)
                else: 
                    specific_results[key] = val

    # Si pas trouvé, mettre 0
    if iterations is None:
        iterations = 0

    return Demande_values, lp_value, cl_value, total_nonzeros, sum_X, sum_Y, sum_B, sum_NAJ, sum_NRE, iterations, specific_results



# --- Fonctions de parsing LTF (inchangé sauf adaptation si nécessaire) ---
def process_ltf(file_path):
    if not os.path.exists(file_path):
        print(f"[ERREUR] Le fichier .ltf n'existe pas : {file_path}")
        return {}, "", "", 0, 0, 0, 0, 0, 0, {}

    sum_DEM = 0
    sum_X = sum_Y = sum_B = sum_NAJ = sum_NRE = 0
    lp_value = ""
    cl_value = ""
    total_nonzeros = 0
    specific_results = {}
    dem_values = {}

    with open(file_path, 'r') as f:
        lines = f.readlines()

    for line in lines:
        # DEM
        match = re.search(r"DEM\(\s*(\d+),\s*(\d+)\)\s+(\d+\.\d+)\s+\d+\.\d+", line)
        if match:
            key = f"d{match.group(1)}{match.group(2)}"
            value = round(float(match.group(3)))
            dem_values[key] = value
            sum_DEM += value

        # Variables X,Y,B,NAJ,NRE,MN
        for var_type, regex in [
            ("X", r"X\(\s*\d+,\s*\d+,\s*\d+,\s*\d+,\s*\d+,\s*\d+\)\s+(\d+\.\d+)"),
            ("Y", r"Y\(\s*\d+,\s*\d+,\s*\d+\)\s+(\d+\.\d+)"),
            ("B", r"B\(\s*\d+,\s*\d+\)\s+(\d+\.\d+)"),
            ("NAJ", r"NAJ\(\s*\d+,\s*\d+,\s*\d+\)\s+(\d+\.\d+)"),
            ("NRE", r"NRE\(\s*\d+,\s*\d+,\s*\d+\)\s+(\d+\.\d+)"),
            ("MN", r"MN\(\s*\d+,\s*\d+,\s*\d+\)\s+(\d+\.\d+)")
        ]:
            m = re.search(regex, line)
            if m:
                val = float(m.group(1))
                if var_type == "X": sum_X += val
                elif var_type == "Y": sum_Y += val
                elif var_type == "B": sum_B += val
                elif var_type == "NAJ": sum_NAJ += val
                elif var_type == "NRE": sum_NRE += val

        # Patterns spécifiques
        patterns = {
            "Objective value": r"Objective value:\s+([-+]?\d+\.\d+)",
            "Objective bound": r"Objective bound:\s+([-+]?\d+\.\d+)",
            "Total variables": r"Total variables:\s+(\d+)",
            "Integer variables": r"Integer variables:\s+(\d+)",
            "Total solver iterations": r"Total solver iterations:\s+(\d+)",
            "Total constraints": r"Total constraints:\s+(\d+)",
            "Elapsed runtime (s)": r"Elapsed runtime seconds:\s+(\d+\.\d+)",
            "Total nonzeros": r"Total nonzeros:\s+(\d+)",
            "LP": r"LP\s+(\d+\.\d+)",
            "CL": r"CL\s+(\d+\.\d+)"
        }

        for key, regex in patterns.items():
            m = re.search(regex, line)
            if m:
                val = m.group(1)
                if key == "Total nonzeros": total_nonzeros = val
                elif key == "LP": lp_value = val
                elif key == "CL": cl_value = val
                else: specific_results[key] = val

    return dem_values, lp_value, cl_value, total_nonzeros, sum_X, sum_Y, sum_B, sum_NAJ, sum_NRE, specific_results

# --- Traitement TXT & LTF ---
Dem_txt, lp_txt, cl_txt, nonzeros_txt, sum_X_txt, sum_Y_txt, sum_B_txt, sum_NAJ_txt, sum_NRE_txt, iterations_txt, res_txt = process_txt(TXT_FILE)
Dem_ltf, lp_ltf, cl_ltf, nonzeros_ltf, sum_X_ltf, sum_Y_ltf, sum_B_ltf, sum_NAJ_ltf, sum_NRE_ltf, res_ltf = process_ltf(LTF_FILE)

# --- Ajout dans Excel ---
data_row_txt = [
    "resultats_optimisation.txt", scenario, instance_scenario,
    Dem_txt.get("d11",0),Dem_txt.get("d12",0),Dem_txt.get("d21",0),Dem_txt.get("d22",0),
    Dem_txt.get("d31",0),Dem_txt.get("d32",0),Dem_txt.get("d41",0),Dem_txt.get("d42",0),
    lp_txt, cl_txt, res_txt.get("Fonction objectif",""),
    sum_X_txt, sum_B_txt, 1, sum_Y_txt, sum_NAJ_txt, sum_NRE_txt,
    res_txt.get("Temps d'exécution",""), nonzeros_txt,
    res_txt.get("Variables totales",""), res_txt.get("Contraintes totales",""),
    res_txt.get("Nombre d'itérations","")
]
ws_txt.append(data_row_txt)

data_row_ltf = [
    "DataFinal.ltf", scenario, instance_scenario,
    Dem_ltf.get("d11",0),Dem_ltf.get("d12",0),Dem_ltf.get("d21",0),Dem_ltf.get("d22",0),
    Dem_ltf.get("d31",0),Dem_ltf.get("d32",0),Dem_ltf.get("d41",0),Dem_ltf.get("d42",0),
    lp_ltf, cl_ltf, sum(Dem_ltf.values()), res_ltf.get("Objective value",""), res_ltf.get("Objective bound",""),
    sum_X_ltf, sum_B_ltf, 1, sum_Y_ltf, sum_NAJ_ltf, sum_NRE_ltf,
    res_ltf.get("Elapsed runtime (s)",""), res_ltf.get("Integer variables",""), nonzeros_ltf,
    res_ltf.get("Total variables",""), res_ltf.get("Total constraints",""), "*", res_ltf.get("Total solver iterations","")
]
ws_ltf.append(data_row_ltf)

# --- 🔹 Synthèse ---
def create_synthese(workbook):
    # Supprimer l'ancienne synthèse si elle existe
    if "Synthese" in workbook.sheetnames:
        del workbook["Synthese"]

    ws_syn = workbook.create_sheet("Synthese")
    ws_syn.append(["Feuille - Indicateur", "Moyenne", "Maximum", "Dernière valeur"])

    mots_cles = ["OBJ", "PROD", "REPORT", "SOUS", "TEMPS", "EXECUTION", "OBJECTIVE", "FONCTION"]

    # --- Parcours des feuilles pour calcul des moyennes / max ---
    for ws in workbook.worksheets:
        if ws.title == "Synthese":
            continue

        col_names = list(ws[1])

        for i, cell in enumerate(col_names, start=1):
            col_vals = []
            if (
                cell.value
                and isinstance(cell.value, str)
                and any(x in cell.value.upper() for x in mots_cles)
            ):
                col_vals = [
                    r[i - 1]
                    for r in ws.iter_rows(min_row=2, values_only=True)
                    if isinstance(r[i - 1], (int, float))
                ]

            if col_vals:
                ws_syn.append([
                    f"{ws.title} - {cell.value}",
                    round(statistics.mean(col_vals), 3) if len(col_vals) > 1 else col_vals[0],
                    round(max(col_vals), 3),
                    round(col_vals[-1], 3)
                ])

    # --- 🔹 Ajout explicite des valeurs de fonction objectif, Iterations et Process Time ---
    try:
        # TXT
        if "TXT_Resultats" in workbook.sheetnames:
            ws_txt = workbook["TXT_Resultats"]
            last_row_txt = [cell.value for cell in list(ws_txt.rows)[-1]]
            
            # Fonction objectif
            idx_obj_txt = headers_txt.index("OBJ.VALUE")
            val_obj_txt = last_row_txt[idx_obj_txt]
            ws_syn.append(["TXT_Resultats - Fonction objectif", "", "", val_obj_txt])
            
            # Iterations
            idx_iter_txt = headers_txt.index("ITERATIONS")
            val_iter_txt = last_row_txt[idx_iter_txt]
            ws_syn.append(["TXT_Resultats - Iterations", "", "", val_iter_txt])
            
            # Process Time
            idx_time_txt = headers_txt.index("PRECESS TIME")
            val_time_txt = last_row_txt[idx_time_txt]
            ws_syn.append(["TXT_Resultats - Process Time (s)", "", "", val_time_txt])

        # LTF
        if "LTF_Resultats" in workbook.sheetnames:
            ws_ltf = workbook["LTF_Resultats"]
            last_row_ltf = [cell.value for cell in list(ws_ltf.rows)[-1]]
            
            # Fonction objectif
            idx_obj_ltf = headers_ltf.index("OBJ.VALUE")
            val_obj_ltf = last_row_ltf[idx_obj_ltf]
            ws_syn.append(["LTF_Resultats - Objective value", "", "", val_obj_ltf])
            
            # Iterations
            idx_iter_ltf = headers_ltf.index("ITERATIONS")
            val_iter_ltf = last_row_ltf[idx_iter_ltf]
            ws_syn.append(["LTF_Resultats - Iterations", "", "", val_iter_ltf])
            
            # Process Time
            idx_time_ltf = headers_ltf.index("PRECESS TIME")
            val_time_ltf = last_row_ltf[idx_time_ltf]
            ws_syn.append(["LTF_Resultats - Process Time (s)", "", "", val_time_ltf])

    except Exception as e:
        print(f"[AVERTISSEMENT] Impossible d’ajouter les objectifs, Iterations ou Process Time : {e}")

    # --- Ajustement largeur colonnes ---
    for col in ws_syn.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws_syn.column_dimensions[col_letter].width = max_length + 2

    return ws_syn





create_synthese(wb)

# --- Sauvegarde ---
wb.save(OUTPUT_EXCEL)
print(f"[OK] Résultats enregistrés avec synthèse dans {OUTPUT_EXCEL}")


